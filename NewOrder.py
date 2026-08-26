import os
import time
import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException, ElementClickInterceptedException
from openpyxl import load_workbook, Workbook

# Импортируем конфигурацию
from config import config

class KDVApp:
    def __init__(self):
        
        self.driver = None
        self.wait = None

        # Используем данные из конфигурации
        self.email = config.KDV_EMAIL
        self.password = config.KDV_PASSWORD
        self.excel_input = config.EXCEL_INPUT_PATH
        self.excel_output = config.EXCEL_OUTPUT_PATH
        self.max_products = config.MAX_PRODUCTS_TO_PROCESS
        
        print(f"📧 Email: {self.email}")
        print(f"📊 Максимум товаров: {self.max_products}")
        print(f"📁 Входной файл: {self.excel_input}")
        print(f"📁 Выходной файл: {self.excel_output}")
        
    def setup_driver(self):
        """Настройка Chrome драйвера"""
        chrome_options = Options()
        
        prefs = {
            "profile.managed_default_content_settings.images": 2,
            "profile.managed_default_content_settings.stylesheets": 2,
            "profile.managed_default_content_settings.javascript": 1
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        chrome_options.add_argument("--incognito")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--remote-allow-origins=*")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36")

        
         # Headless режим из конфигурации
        if config.HEADLESS_MODE:
            chrome_options.add_argument("--headless")
            print("🧪 Headless режим включен")

        # Отключаем логи
        chrome_options.add_argument("--log-level=3")  # 0=INFO, 1=WARNING, 2=ERROR, 3=FATAL
        chrome_options.add_argument("--silent")
        
        # Отключаем GPU и другие сообщения
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-software-rasterizer")
        
        # Скрываем DevTools сообщения
        chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
            
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        self.wait = WebDriverWait(self.driver, 15)
        
    def check_is_exit(self, result):
        """Проверка на выход из программы"""
        if result.lower() in ["выход", "exit"]:
            print("Завершаю работу")
            sys.exit(0)
            
    def validate_user_input(self, user_input):
        """Валидация ввода пользователя"""
        if not user_input or len(user_input) != 1:
            raise ValueError(f"Incorrect user input: expected one digit as answer, but actually get {user_input}")
        
        try:
            answer = int(user_input)
            if answer not in [1, 2, 3]:
                raise ValueError(f"Incorrect user input: expected 1 or 2 or 3, but actually get {user_input}")
            return answer
        except ValueError:
            raise ValueError("Incorrect user input: character is not numeric!")
            
    def read_excel_file(self, file_path):
        """Чтение Excel файла"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл {file_path} не найден!")
        
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        print(f"📊 Чтение Excel файла: {file_path}")
        print(f"📊 Всего строк в файле: {sheet.max_row}")
        print(f"📊 Всего колонок в файле: {sheet.max_column}")
        
        data = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)
        
        print(f"📊 Найдено непустых строк: {len(data)}")
        
        if data:
            print("\n📋 Первые 5 строк данных:")
            for i, row in enumerate(data[:5], start=1):
                print(f"  Строка {i}: {row}")
        
        return data, len(data), sheet.max_column
    
    def save_excel_file(self, file_path, data, max_col):
        """Сохранение данных в Excel"""
        workbook = Workbook()
        sheet = workbook.active
        
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                if value is not None:
                    sheet.cell(row=row_idx, column=col_idx, value=value)
        
        workbook.save(file_path)
        print(f"✅ Результат сохранен в {file_path}")
    
    def login_to_kdv(self):
        """Авторизация на KDV через форму"""
        print("\n🔐 Авторизация на KDV...")
        
        self.driver.get("https://kdvonline.ru/signinlegal")
        time.sleep(2)
        
        try:
            # Ждем загрузки формы
            self.wait.until(EC.presence_of_element_located((By.TAG_NAME, "form")))
            
            # Ищем поле для email
            email_input = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='identity']"))
            )
            email_input.clear()
            email_input.send_keys(self.email)
            print(f"✅ Введен email: {self.email}")
            
            # Ищем поле для пароля
            password_input = self.driver.find_element(By.CSS_SELECTOR, "input[name='credential']")
            password_input.clear()
            password_input.send_keys(self.password)
            print(f"✅ Введен пароль")
            
            # Ищем кнопку "Войти"
            login_button = self.driver.find_element(By.XPATH, "//button[@type='submit']//span[contains(text(), 'Войти')]")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", login_button)
            time.sleep(0.5)
            
            try:
                login_button.click()
            except ElementClickInterceptedException:
                self.driver.execute_script("arguments[0].click();", login_button)
            
            print("⏳ Ожидание авторизации...")
            time.sleep(3)
            
            # Проверяем успешность авторизации
            try:
                self.wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".acw46Ay17, .user-menu, .profile-link"))
                )
                print("✅ Авторизация успешна!")
                return True
            except TimeoutException:
                print("⚠️ Не удалось подтвердить авторизацию, но продолжаем...")
                return True
                
        except Exception as e:
            print(f"❌ Ошибка при авторизации: {str(e)}")
            return False
    
    def wait_for_cart_update(self, timeout=10):
        """
        Ожидание обновления корзины после добавления товара
        """
        start_time = time.time()
        
        # Ждем появления индикатора загрузки и его исчезновения
        try:
            # Ждем появления индикатора загрузки (если есть)
            self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".loading, .spinner, .loader, .cart-loading"))
            )
            print("⏳ Обнаружен процесс добавления в корзину...")
            
            # Ждем исчезновения индикатора загрузки
            self.wait.until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, ".loading, .spinner, .loader, .cart-loading"))
            )
            print("✅ Добавление в корзину завершено")
            return True
        except TimeoutException:
            # Если индикатора загрузки нет, проверяем изменения в DOM
            pass
        
        # Альтернативный метод: ждем обновления счетчика корзины
        try:
            # Сохраняем текущее количество товаров в корзине
            try:
                cart_count_before = self.driver.find_element(By.CSS_SELECTOR, ".cart-count, .cart-items-count, .basket-count")
                count_before = cart_count_before.text
            except:
                count_before = None
            
            # Ждем изменения счетчика
            for i in range(timeout * 2):  # Проверяем каждые 0.5 секунды
                time.sleep(0.5)
                try:
                    cart_count_after = self.driver.find_element(By.CSS_SELECTOR, ".cart-count, .cart-items-count, .basket-count")
                    count_after = cart_count_after.text
                    if count_before != count_after:
                        print(f"✅ Корзина обновлена: {count_after} товаров")
                        return True
                except:
                    pass
            
            print("✅ Продолжаем (корзина обновлена)")
            return True
        except Exception as e:
            print(f"⚠️ Не удалось отследить обновление корзины: {e}")
            return True
    
    def find_add_to_cart_button(self):
        """Поиск кнопки 'В корзину'"""
        selectors = [
            #(By.XPATH, 'value': "//div[.//span[contains(text(), '₽')] and .//button[contains(text(), 'В корзину')]]"(),
            #(By.CSS_SELECTOR, "span.qoFy5xub4.cKAwCiQ37.XoFy5xub4.aoFy5xub4"),
            #(By.CSS_SELECTOR, "span[class*='qoFy5xub4']"),
            #(By.CSS_SELECTOR, "span[class*='cKAwCiQ37']"),
            #(By.XPATH, "//span[contains(text(), 'В корзину')]"),
            (By.XPATH, "//span[contains(text(), 'В корзину') and parent::*//span[contains(text(), '₽')]]")
            #(By.XPATH, "//*[contains(text(), 'В корзину') and not(contains(text(), 'Нет'))]"),
        ]
        
        for by, selector in selectors:
            try:
                elements = self.driver.find_elements(by, selector)
                for element in elements:
                    if element.is_displayed() and element.is_enabled():
                        print(f"✅ Найдена кнопка по селектору: {selector}")
                        return element
            except Exception:
                continue
        
        return None
    
    def process_kdv_orders(self):
        """Обработка заказов KDV"""
        print("\n🔄 Начинаю обработку заказов KDV...")
        
        # Чтение Excel файла
        file_path = "C:/TEMP/read.xlsx"
        try:
            data, max_row, max_col = self.read_excel_file(file_path)
        except FileNotFoundError as e:
            print(f"❌ {e}")
            return
        
        if not data or max_row == 0:
            print("❌ Файл Excel пуст или не содержит данных!")
            return
                
        # Авторизация
        if not self.login_to_kdv():
            print("❌ Не удалось авторизоваться. Проверьте логин и пароль.")
            return
        
        new_data = []
        
        # Определяем колонку с URL
        url_column_index = None
        for row in data:
            for col_idx, value in enumerate(row):
                if value and isinstance(value, str) and "https://kdvonline.ru/" in value:
                    url_column_index = col_idx
                    break
            if url_column_index is not None:
                break
        
        if url_column_index is None:
            print("❌ Не найдено ни одного URL в файле!")
            return
        
        print(f"\n🔍 URL-адреса найдены в колонке {url_column_index + 1}")
        
        # Обработка товаров
        processed_count = 0
        total_products = 0
        
        # Подсчет общего количества товаров для обработки
        for row_data in data:
            if len(row_data) > url_column_index and row_data[url_column_index]:
                if "https://kdvonline.ru/" in str(row_data[url_column_index]):
                    total_products += 1
        
        print(f"📊 Всего товаров для обработки: {total_products}")
        
        for row_idx, row_data in enumerate(data, start=1):
        # 🔥 ПРОВЕРКА В НАЧАЛЕ ЦИКЛА
            if processed_count >= self.max_products:
                print(f"\n⏹️ Достигнут лимит в {self.max_products} товаров")
                # Добавляем оставшиеся строки без обработки
                for remaining_row in data[row_idx-1:]:
                    new_data.append(list(remaining_row) if remaining_row else [])
                break

            new_row = list(row_data) if row_data else []
            
            if len(new_row) > url_column_index and new_row[url_column_index]:
                product_url = str(new_row[url_column_index])
                
                if "https://kdvonline.ru/" in product_url:

                    
                    print(f"\n🛒 [{processed_count + 1}/{total_products}] Обработка товара: {product_url}")
                    
                    try:
                        self.driver.get(product_url)
                        processed_count += 1
                        time.sleep(2)  # Уменьшено время ожидания
                        
                        # Проверяем доступность товара
                        try:
                            unavailable_selectors = [
                                (By.XPATH, "//*[contains(text(), 'Нет в наличии')]"),
                                (By.XPATH, "//*[contains(text(), 'Товар недоступен')]"),
                                (By.XPATH, "//*[contains(text(), 'Нет на складе')]"),
                                (By.XPATH, "//*[contains(text(), 'Уведомить о поступлении')]"),
                                (By.XPATH, "//*[contains(text(), 'Отписаться')]")
                            ]
                            
                            for by, selector in unavailable_selectors:
                                try:
                                    if self.driver.find_element(by, selector).is_displayed():
                                        print(f"⚠️ Товар недоступен")
                                        if len(new_row) <= 8:
                                            new_row.extend([None] * (9 - len(new_row)))
                                        new_row[8] = "Нет в наличии"
                                        new_data.append(new_row)
                                        raise Exception("Товар недоступен")
                                except NoSuchElementException:
                                    continue
                        except Exception as e:
                            if "Товар недоступен" in str(e):
                                continue
                        
                        # Ищем кнопку "В корзину"
                        add_button = self.find_add_to_cart_button()
                        
                        if add_button:
                            # Прокручиваем к кнопке
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", add_button)
                            time.sleep(0.5)
                            
                            # Проверяем количество
                            quantity_value = None
                            if len(new_row) > 3 and new_row[3] and isinstance(new_row[3], (int, float)):
                                quantity_value = new_row[3]
                            elif len(new_row) > 4 and new_row[4] and isinstance(new_row[4], (int, float)):
                                quantity_value = new_row[4]                         

                            
                            # Кликаем по кнопке
                            try:
                                add_button.click()
                                time.sleep(0.5)
                                if quantity_value and quantity_value > 1:
                                    #clicks_needed = int(quantity_value) - 1
                                    #for i in range(clicks_needed):
                                        try:
                                            # Ищем поле ввода
                                            quantity_input = self.driver.find_element(By.CSS_SELECTOR, "input.dzbJpVxc5")
                                            
                                            # Очищаем поле
                                            quantity_input.clear()
                                            
                                            # Вводим новое значение
                                            quantity_input.send_keys(Keys.BACKSPACE)
                                            quantity_input.send_keys(str(quantity_value))
                                            
                                            # Убираем фокус с поля (чтобы обновить цену)
                                            quantity_input.send_keys(Keys.TAB)
                                            
                                            print(f"✅ Количество установлено: {quantity_value}")
                                            
                                            time.sleep(0.5)  # Небольшая пауза между нажатиями
                                            
                                        except Exception as e:
                                            print(f"  ❌ Ошибка при нажатии: {e}")
                                            return False
                                    
                                        #print(f"✅ Количество установлено: {quantity_value}")
                                        
                                        
                                print("✅ Клик по кнопке выполнен")
                            except ElementClickInterceptedException:
                                self.driver.execute_script("arguments[0].click();", add_button)
                                print("✅ Клик через JavaScript выполнен")
                            
                            # Ждем обновления корзины (уменьшено время ожидания)
                            #self.wait_for_cart_update(timeout=5)
                            
                            # Записываем результат
                            if len(new_row) <= 5:
                                new_row.extend([None] * (6 - len(new_row)))
                            new_row[5] = quantity_value if quantity_value else 1
                            
                            print(f"✅ Товар добавлен в корзину")
                            
                        else:
                            print(f"⚠️ Не найдена кнопка 'В корзину'")
                            # Проверяем кнопку подписки
                            try:
                                subscription = self.driver.find_element(
                                    By.XPATH, "//*[contains(text(), 'Подписаться') or contains(text(), 'Уведомить о поступлении')]"
                                )
                                if subscription.is_displayed():
                                    subscription.click()
                                    print("✅ Подписка оформлена")
                                    if len(new_row) <= 8:
                                        new_row.extend([None] * (9 - len(new_row)))
                                    new_row[8] = "Нет в наличии"
                            except NoSuchElementException:
                                print(f"❌ Не найдена кнопка действия")
                                
                    except Exception as e:
                        if "Товар недоступен" not in str(e):
                            print(f"❌ Ошибка при обработке: {str(e)}")
            
            new_data.append(new_row)
        
        # Сохраняем результат
        output_path = "C:/TEMP/read_new.xlsx"
        self.save_excel_file(output_path, new_data, max_col + 5)
        print(f"\n📊 Обработано товаров: {processed_count} из {total_products}")
        print(f"📊 Всего строк в файле: {len(new_data)}")
        
        # Переход в корзину
        self.driver.get("https://kdvonline.ru/profile/orders")
    
    def run(self):
        """Основной метод запуска"""
        try:
            # Проверяем конфигурацию
            config.validate()
        
            while True:
                print("\n" + "="*50)
                print("Введите ответ:")
                print("1 - Заказ KDV")
                print("2 - Заказ Happywear")
                print("выход (exit) - завершить работу")
                print("="*50)
                
                result = input().strip()
                self.check_is_exit(result)
                
                try:
                    answer = self.validate_user_input(result)
                    break
                except ValueError as e:
                    print(f"❌ Ошибка: {e}")
                    continue
            
            # Настраиваем драйвер
            print("\n🚀 Запуск браузера...")
            self.setup_driver()
            
            # Выполняем действие
            if answer == 1:
                self.process_kdv_orders()
            elif answer == 2:
                print("Функционал Happywear требует отдельной настройки")
            else:
                print("Неизвестный выбор")
                
        except Exception as e:
            print(f"❌ Критическая ошибка: {str(e)}")
            import traceback
            traceback.print_exc()
            
        finally:
            if self.driver:
                time.sleep(2)
                self.driver.quit()
                print("\n🔚 Драйвер закрыт")

if __name__ == "__main__":
    app = KDVApp()
    app.run()